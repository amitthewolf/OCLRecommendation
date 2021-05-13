import os
from openpyxl import load_workbook
import openpyxl as op
import pandas as pd
from sklearn.metrics import accuracy_score
from datetime import datetime

class Logger:

    def __init__(self):
        self.filename= '../Outputs/Results.xlsx'

    def LogResult(self,modelName, YTest, PredTest, YTrain, PredTrain, scores, test_config, featureNames, iterations, fixed_section):
        if test_config.n2v_flag == 'True':
            data = {'Timestamp': datetime.now(),
                    'Target': test_config.target,
                    'Method': test_config.method,
                    'Model': modelName,
                    'Features': ','.join(featureNames[:-1]),
                    'Sampling': test_config.sampling_strategy,
                    'Test_ratio': test_config.test_ratio,
                    'Cross_Val_K': test_config.cross_val_k,
                    'Graphlets': fixed_section['graphlets_flag'],
                    'N2V_Flag': test_config.n2v_flag,
                    'N2V_Features_Num': test_config.n2v_features_num,
                    'N2V_use_Att': 'true',
                    'N2V_use_Inhe': 'true',
                    'N2V_ReturnWe': test_config.n2v_return_weight,
                    'N2V_WalkLen': test_config.n2v_walklen,
                    'N2V_Epochs': test_config.n2v_epochs,
                    'N2V_NeighborWeight': test_config.n2v_neighbor_weight,
                    'Num_PCA': test_config.pca,
                    'Iterations': iterations,
                    'Random': 'null',
                    'Train Score': accuracy_score(YTrain, PredTrain),
                    'Test Score': accuracy_score(YTest, PredTrain),
                    'Mean': scores.mean(),
                    'Std': scores.std(),
                    }

            Log_DF = pd.DataFrame(data, columns=['Timestamp',
                                                 'Target',
                                                 'Method',
                                                 'Model',
                                                 'Features',
                                                 'Sampling',
                                                 'Test_ratio',
                                                 'Cross_Val_K',
                                                 'Graphlets',
                                                 'N2V_Flag',
                                                 'N2V_Features_Num',
                                                 'N2V_use_Att',
                                                 'N2V_use_Inhe',
                                                 'N2V_ReturnWe',
                                                 'N2V_WalkLen',
                                                 'N2V_Epochs',
                                                 'N2V_NeighborWeight',
                                                 'Num_PCA',
                                                 'Iterations',
                                                 'Random',
                                                 'Train Score',
                                                 'Test Score',
                                                 'Mean',
                                                 'Std'], index=[0])
        else:
            data = {'Timestamp': datetime.now(),
                    'Target': test_config.target,
                    'Method': test_config.method,
                    'Model': modelName,
                    'Features': ','.join(featureNames[:-1]),
                    'Sampling': test_config.sampling_strategy,
                    'Test_ratio': test_config.test_ratio,
                    'Cross_Val_K': test_config.cross_val_k,
                    'Graphlets': fixed_section['graphlets_flag'],
                    'N2V_Flag': test_config.n2v_flag,
                    'N2V_Features_Num': '-',
                    'N2V_use_Att': '-',
                    'N2V_use_Inhe': '-',
                    'N2V_ReturnWe': '-',
                    'N2V_WalkLen': '-',
                    'N2V_Epochs': '-',
                    'N2V_NeighborWeight': '-',
                    'Num_PCA': '-',
                    'Iterations': iterations,
                    'Random': 'null',
                    'Train Score': accuracy_score(YTrain, PredTrain),
                    'Test Score': accuracy_score(YTest, PredTest),
                    'Mean': scores.mean(),
                    'Std': scores.std(),
                    }

            Log_DF = pd.DataFrame(data, columns=['Timestamp',
                                                 'Target',
                                                 'Method',
                                                 'Model',
                                                 'Features',
                                                 'Sampling',
                                                 'Test_ratio',
                                                 'Cross_Val_K',
                                                 'Graphlets',
                                                 'N2V_Flag',
                                                 'N2V_Features_Num',
                                                 'N2V_use_Att',
                                                 'N2V_use_Inhe',
                                                 'N2V_ReturnWe',
                                                 'N2V_WalkLen',
                                                 'N2V_Epochs',
                                                 'N2V_NeighborWeight',
                                                 'Num_PCA',
                                                 'Iterations',
                                                 'Random',
                                                 'Train Score',
                                                 'Test Score',
                                                 'Mean',
                                                 'Std'], index=[0])
        # log = Logger()
        self.append_df_to_excel(Log_DF, header=None, index=False)

    def LogResultOperator(self, modelName, Train, Test, test_config, featureNames, iterations, fixed_section):
        if test_config.n2v_flag == 'True':
            data = {'Timestamp': datetime.now(),
                    'Target': test_config.target,
                    'Method': test_config.method,
                    'Model': modelName,
                    'Features': ','.join(featureNames[:-1]),
                    'Sampling': test_config.sampling_strategy,
                    'Test_ratio': test_config.test_ratio,
                    'Cross_Val_K': test_config.cross_val_k,
                    'Graphlets': fixed_section['graphlets_flag'],
                    'N2V_Flag': test_config.n2v_flag,
                    'N2V_Features_Num': test_config.n2v_features_num,
                    'N2V_use_Att': 'true',
                    'N2V_use_Inhe': 'true',
                    'N2V_ReturnWe': test_config.n2v_return_weight,
                    'N2V_WalkLen': test_config.n2v_walklen,
                    'N2V_Epochs': test_config.n2v_epochs,
                    'N2V_NeighborWeight': test_config.n2v_neighbor_weight,
                    'Num_PCA': test_config.pca,
                    'Iterations': iterations,
                    'Random': 'null',
                    'Train Score': Train,
                    'Test Score': Test,
                    'Mean': '-',
                    'Std': '-',
                    }

            Log_DF = pd.DataFrame(data, columns=['Timestamp',
                                                 'Target',
                                                 'Method',
                                                 'Model',
                                                 'Features',
                                                 'Sampling',
                                                 'Test_ratio',
                                                 'Cross_Val_K',
                                                 'Graphlets',
                                                 'N2V_Flag',
                                                 'N2V_Features_Num',
                                                 'N2V_use_Att',
                                                 'N2V_use_Inhe',
                                                 'N2V_ReturnWe',
                                                 'N2V_WalkLen',
                                                 'N2V_Epochs',
                                                 'N2V_NeighborWeight',
                                                 'Num_PCA',
                                                 'Iterations',
                                                 'Random',
                                                 'Train Score',
                                                 'Test Score',
                                                 'Mean',
                                                 'Std'], index=[0])
        else:
            data = {'Timestamp': datetime.now(),
                    'Target': test_config.target,
                    'Method': test_config.method,
                    'Model': modelName,
                    'Features': ','.join(featureNames[:-1]),
                    'Sampling': test_config.sampling_strategy,
                    'Test_ratio': test_config.test_ratio,
                    'Cross_Val_K': test_config.cross_val_k,
                    'Graphlets': fixed_section['graphlets_flag'],
                    'N2V_Flag': test_config.n2v_flag,
                    'N2V_Features_Num': '-',
                    'N2V_use_Att': '-',
                    'N2V_use_Inhe': '-',
                    'N2V_ReturnWe': '-',
                    'N2V_WalkLen': '-',
                    'N2V_Epochs': '-',
                    'N2V_NeighborWeight': '-',
                    'Num_PCA': '-',
                    'Iterations': iterations,
                    'Random': 'null',
                    'Train Score': Train,
                    'Test Score': Test,
                    'Mean': '-',
                    'Std': '-',
                    }

            Log_DF = pd.DataFrame(data, columns=['Timestamp',
                                                 'Target',
                                                 'Method',
                                                 'Model',
                                                 'Features',
                                                 'Sampling',
                                                 'Test_ratio',
                                                 'Cross_Val_K',
                                                 'Graphlets',
                                                 'N2V_Flag',
                                                 'N2V_Features_Num',
                                                 'N2V_use_Att',
                                                 'N2V_use_Inhe',
                                                 'N2V_ReturnWe',
                                                 'N2V_WalkLen',
                                                 'N2V_Epochs',
                                                 'N2V_NeighborWeight',
                                                 'Num_PCA',
                                                 'Iterations',
                                                 'Random',
                                                 'Train Score',
                                                 'Test Score',
                                                 'Mean',
                                                 'Std'], index=[0])
        #log = Logger()
        self.append_df_to_excel(Log_DF, header=None, index=False)

    def append_df_to_excel(self, df, startrow=None, truncate_sheet=False, **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        filename = '../Outputs/Results.xlsx'
        sheet_name = 'Result Table'
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

    def LogSamples(self, df,classify, startrow=None, truncate_sheet=False, **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        filename = '../Outputs/Samples.xlsx'
        if classify == 'FP':
            sheet_name = 'False Positives'
        elif classify == 'FN':
            sheet_name = 'False Negatives'

        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)
        if 'False Positives' not in writer.book.sheetnames:
            writer.book.create_sheet('False Positives')
        if 'False Negatives' not in writer.book.sheetnames:
            writer.book.create_sheet('False Negatives')

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()
